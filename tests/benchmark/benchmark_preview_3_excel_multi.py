#!/usr/bin/env python3
# Generate Excel from multiple run JSON files with statistics

import json
import sys
from datetime import datetime
from pathlib import Path
import statistics

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill
except ImportError:
    print("❌ openpyxl not installed. Install with: pip3 install openpyxl")
    sys.exit(1)

def load_runs(base_dir, test_type, num_runs=5):
    """Load all JSON files for a test type"""
    runs = []
    for i in range(1, num_runs + 1):
        try:
            filepath = base_dir / f'{test_type}_run_{i}.json'
            with open(filepath, 'r') as f:
                runs.append(json.load(f))
        except FileNotFoundError:
            print(f"Warning: {filepath} not found")
    return runs

def calculate_test_stats(runs):
    """Calculate statistics for each test across runs"""
    # Collect all tests by fullTitle
    test_data = {}
    
    for run in runs:
        for test in run['tests']:
            full_title = test['fullTitle']
            if full_title not in test_data:
                test_data[full_title] = {
                    'durations': [],
                    'speed': test.get('speed', 'n/a'),
                    'state': test.get('state', 'unknown'),
                    'test': test
                }
            test_data[full_title]['durations'].append(test.get('duration', 0))
    
    # Calculate stats for each test
    results = []
    for full_title, data in sorted(test_data.items()):
        durations = data['durations']
        results.append({
            'fullTitle': full_title,
            'test': data['test'],
            'runs': durations,
            'average': statistics.mean(durations) if durations else 0,
            'stddev': statistics.stdev(durations) if len(durations) > 1 else 0,
            'min': min(durations) if durations else 0,
            'max': max(durations) if durations else 0,
            'speed': data['speed'],
            'state': data['state']
        })
    
    return results

def calculate_suite_stats(runs):
    """Calculate overall statistics for the test suite"""
    durations = [run['stats']['duration'] for run in runs]
    return {
        'tests': runs[0]['stats']['tests'],
        'passes': runs[0]['stats']['passes'],
        'failures': runs[0]['stats']['failures'],
        'suites': runs[0]['stats']['suites'],
        'runs': durations,
        'average': statistics.mean(durations),
        'stddev': statistics.stdev(durations) if len(durations) > 1 else 0,
        'min': min(durations),
        'max': max(durations)
    }

def main():
    base_dir = Path(__file__).parent
    runs_dir = base_dir / 'benchmark_results' / 'runs'
    
    print("Loading test results...")
    mocha_runs = load_runs(runs_dir, 'mocha')
    plunit_runs = load_runs(runs_dir, 'plunit')
    
    if not mocha_runs or not plunit_runs:
        print("❌ Error: No run data found")
        sys.exit(1)
    
    print(f"Found {len(mocha_runs)} Mocha runs and {len(plunit_runs)} PL-Unit runs")
    print("Calculating statistics...")
    
    mocha_stats = calculate_suite_stats(mocha_runs)
    plunit_stats = calculate_suite_stats(plunit_runs)
    
    mocha_tests = calculate_test_stats(mocha_runs)
    plunit_tests = calculate_test_stats(plunit_runs)
    
    print("Generating Excel report...")
    
    # Create workbook
    wb = openpyxl.Workbook()
    
    # Tab 1: Summary
    ws_summary = wb.active
    ws_summary.title = "Summary"
    
    # Headers
    num_runs = len(mocha_runs)
    headers = ['Test Suite', 'Tests', 'Passes', 'Failures', 'Suites']
    for i in range(1, num_runs + 1):
        headers.append(f'Run {i} (ms)')
    headers.extend(['Average (ms)', 'Std Dev (ms)', 'Min (ms)', 'Max (ms)'])
    
    for col, header in enumerate(headers, 1):
        cell = ws_summary.cell(1, col, header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color='DDDDDD', end_color='DDDDDD', fill_type='solid')
    
    # Mocha row
    ws_summary.cell(2, 1, 'Mocha (capabilities.js)')
    ws_summary.cell(2, 2, mocha_stats['tests'])
    ws_summary.cell(2, 3, mocha_stats['passes'])
    ws_summary.cell(2, 4, mocha_stats['failures'])
    ws_summary.cell(2, 5, mocha_stats['suites'])
    for i, duration in enumerate(mocha_stats['runs'], 6):
        ws_summary.cell(2, i, round(duration))
    col = 6 + num_runs
    ws_summary.cell(2, col, round(mocha_stats['average']))
    ws_summary.cell(2, col + 1, round(mocha_stats['stddev'], 2))
    ws_summary.cell(2, col + 2, round(mocha_stats['min']))
    ws_summary.cell(2, col + 3, round(mocha_stats['max']))
    
    # PL-Unit row
    ws_summary.cell(3, 1, 'PL-Unit (sample_tests)')
    ws_summary.cell(3, 2, plunit_stats['tests'])
    ws_summary.cell(3, 3, plunit_stats['passes'])
    ws_summary.cell(3, 4, plunit_stats['failures'])
    ws_summary.cell(3, 5, plunit_stats['suites'])
    for i, duration in enumerate(plunit_stats['runs'], 6):
        ws_summary.cell(3, i, round(duration))
    col = 6 + num_runs
    ws_summary.cell(3, col, round(plunit_stats['average']))
    ws_summary.cell(3, col + 1, round(plunit_stats['stddev'], 2))
    ws_summary.cell(3, col + 2, round(plunit_stats['min']))
    ws_summary.cell(3, col + 3, round(plunit_stats['max']))
    
    # Adjust column widths
    ws_summary.column_dimensions['A'].width = 30
    for col in ['B', 'C', 'D', 'E']:
        ws_summary.column_dimensions[col].width = 10
    for i in range(6, 6 + num_runs + 4):
        ws_summary.column_dimensions[openpyxl.utils.get_column_letter(i)].width = 12
    
    # Tab 2: Mocha Tests
    ws_mocha = wb.create_sheet("Mocha Tests")
    headers = ['Suite', 'Test Name']
    for i in range(1, num_runs + 1):
        headers.append(f'Run {i} (ms)')
    headers.extend(['Average (ms)', 'Std Dev (ms)', 'Min (ms)', 'Max (ms)', 'Speed', 'Status'])
    
    for col, header in enumerate(headers, 1):
        cell = ws_mocha.cell(1, col, header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color='DDDDDD', end_color='DDDDDD', fill_type='solid')
    
    for row, test_stat in enumerate(mocha_tests, 2):
        full_title = test_stat['fullTitle']
        parts = full_title.split(' ')
        suite = parts[0] if parts else ''
        test_name = ' '.join(parts[1:]) if len(parts) > 1 else full_title
        
        ws_mocha.cell(row, 1, suite)
        ws_mocha.cell(row, 2, test_name)
        
        col = 3
        for duration in test_stat['runs']:
            ws_mocha.cell(row, col, round(duration))
            col += 1
        
        ws_mocha.cell(row, col, round(test_stat['average']))
        ws_mocha.cell(row, col + 1, round(test_stat['stddev'], 2))
        ws_mocha.cell(row, col + 2, round(test_stat['min']))
        ws_mocha.cell(row, col + 3, round(test_stat['max']))
        ws_mocha.cell(row, col + 4, test_stat['speed'])
        ws_mocha.cell(row, col + 5, test_stat['state'])
    
    ws_mocha.column_dimensions['A'].width = 20
    ws_mocha.column_dimensions['B'].width = 50
    for i in range(3, 3 + num_runs + 6):
        ws_mocha.column_dimensions[openpyxl.utils.get_column_letter(i)].width = 12
    
    # Tab 3: PL-Unit Tests
    ws_plunit = wb.create_sheet("PL-Unit Tests")
    for col, header in enumerate(headers, 1):
        cell = ws_plunit.cell(1, col, header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color='DDDDDD', end_color='DDDDDD', fill_type='solid')
    
    for row, test_stat in enumerate(plunit_tests, 2):
        full_title = test_stat['fullTitle']
        if ':' in full_title:
            suite, test_name = full_title.split(':', 1)
        else:
            suite = ''
            test_name = full_title
        
        ws_plunit.cell(row, 1, suite)
        ws_plunit.cell(row, 2, test_name)
        
        col = 3
        for duration in test_stat['runs']:
            ws_plunit.cell(row, col, round(duration))
            col += 1
        
        ws_plunit.cell(row, col, round(test_stat['average']))
        ws_plunit.cell(row, col + 1, round(test_stat['stddev'], 2))
        ws_plunit.cell(row, col + 2, round(test_stat['min']))
        ws_plunit.cell(row, col + 3, round(test_stat['max']))
        ws_plunit.cell(row, col + 4, test_stat['speed'])
        ws_plunit.cell(row, col + 5, test_stat['state'])
    
    ws_plunit.column_dimensions['A'].width = 20
    ws_plunit.column_dimensions['B'].width = 50
    for i in range(3, 3 + num_runs + 6):
        ws_plunit.column_dimensions[openpyxl.utils.get_column_letter(i)].width = 12
    
    # Save file
    timestamp = datetime.now().strftime('%Y-%m-%d_%H%M%S')
    filename = f'benchmark_preview_5runs_{timestamp}.xlsx'
    filepath = base_dir / 'benchmark_results' / filename
    
    wb.save(filepath)
    
    print(f"✅ Excel report generated: {filepath}")
    print(f"\nSummary:")
    print(f"  Mocha:   {mocha_stats['tests']} tests")
    print(f"    Average: {mocha_stats['average']:.0f}ms ± {mocha_stats['stddev']:.1f}ms")
    print(f"    Range: {mocha_stats['min']:.0f}ms - {mocha_stats['max']:.0f}ms")
    print(f"  PL-Unit: {plunit_stats['tests']} tests")
    print(f"    Average: {plunit_stats['average']:.0f}ms ± {plunit_stats['stddev']:.1f}ms")
    print(f"    Range: {plunit_stats['min']:.0f}ms - {plunit_stats['max']:.0f}ms")

if __name__ == '__main__':
    main()
