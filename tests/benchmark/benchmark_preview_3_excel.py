#!/usr/bin/env python3
# Step 3: Generate Excel from JSON files

import json
import sys
from datetime import datetime
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill
except ImportError:
    print("❌ openpyxl not installed. Install with: pip3 install openpyxl")
    sys.exit(1)

def load_json(filename):
    with open(filename, 'r') as f:
        return json.load(f)

def main():
    base_dir = Path(__file__).parent
    results_dir = base_dir / 'benchmark_results'
    
    print("Loading test results...")
    try:
        mocha = load_json(results_dir / 'preview_mocha.json')
        plunit = load_json(results_dir / 'preview_plunit.json')
    except FileNotFoundError as e:
        print(f"❌ Error: {e}")
        print("   Run benchmark_preview_1_mocha.sh and benchmark_preview_2_plunit.sh first")
        sys.exit(1)
    
    print("Generating Excel report...")
    
    # Create workbook
    wb = openpyxl.Workbook()
    
    # Tab 1: Summary (with statistics)
    ws_summary = wb.active
    ws_summary.title = "Summary"
    
    # Headers
    headers = ['Test Suite', 'Tests', 'Passes', 'Failures', 'Suites', 
               'Run 1 (ms)', 'Average (ms)', 'Std Dev (ms)', 'Min (ms)', 'Max (ms)']
    for col, header in enumerate(headers, 1):
        cell = ws_summary.cell(1, col, header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color='DDDDDD', end_color='DDDDDD', fill_type='solid')
    
    # Mocha row
    mocha_duration = mocha['stats']['duration']
    ws_summary.cell(2, 1, 'Mocha (capabilities.js)')
    ws_summary.cell(2, 2, mocha['stats']['tests'])
    ws_summary.cell(2, 3, mocha['stats']['passes'])
    ws_summary.cell(2, 4, mocha['stats']['failures'])
    ws_summary.cell(2, 5, mocha['stats']['suites'])
    ws_summary.cell(2, 6, mocha_duration)
    ws_summary.cell(2, 7, mocha_duration)  # Average = single run for preview
    ws_summary.cell(2, 8, 0)  # Std dev = 0 for single run
    ws_summary.cell(2, 9, mocha_duration)  # Min = duration
    ws_summary.cell(2, 10, mocha_duration)  # Max = duration
    
    # PL-Unit row
    plunit_duration = plunit['stats']['duration']
    ws_summary.cell(3, 1, 'PL-Unit (sample_tests)')
    ws_summary.cell(3, 2, plunit['stats']['tests'])
    ws_summary.cell(3, 3, plunit['stats']['passes'])
    ws_summary.cell(3, 4, plunit['stats']['failures'])
    ws_summary.cell(3, 5, plunit['stats']['suites'])
    ws_summary.cell(3, 6, plunit_duration)
    ws_summary.cell(3, 7, plunit_duration)  # Average = single run for preview
    ws_summary.cell(3, 8, 0)  # Std dev = 0 for single run
    ws_summary.cell(3, 9, plunit_duration)  # Min = duration
    ws_summary.cell(3, 10, plunit_duration)  # Max = duration
    
    # Adjust column widths
    ws_summary.column_dimensions['A'].width = 30
    for col in ['B', 'C', 'D', 'E']:
        ws_summary.column_dimensions[col].width = 10
    for col in ['F', 'G', 'H', 'I', 'J']:
        ws_summary.column_dimensions[col].width = 14
    
    # Tab 2: Mocha Tests (with suite and stats structure)
    ws_mocha = wb.create_sheet("Mocha Tests")
    headers = ['Suite', 'Test Name', 'Run 1 (ms)', 'Average (ms)', 'Std Dev (ms)', 'Speed', 'Status']
    for col, header in enumerate(headers, 1):
        cell = ws_mocha.cell(1, col, header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color='DDDDDD', end_color='DDDDDD', fill_type='solid')
    
    for row, test in enumerate(mocha['tests'], 2):
        full_title = test.get('fullTitle', '')
        # Extract suite name (part before the test name)
        parts = full_title.split(' ')
        suite = parts[0] if parts else ''
        test_name = ' '.join(parts[1:]) if len(parts) > 1 else full_title
        
        duration = test.get('duration', 0)
        
        ws_mocha.cell(row, 1, suite)
        ws_mocha.cell(row, 2, test_name)
        ws_mocha.cell(row, 3, duration)
        ws_mocha.cell(row, 4, duration)  # Average = single run for preview
        ws_mocha.cell(row, 5, 0)  # Std dev = 0 for single run
        ws_mocha.cell(row, 6, test.get('speed', 'n/a'))
        ws_mocha.cell(row, 7, test.get('state', 'unknown'))
    
    ws_mocha.column_dimensions['A'].width = 20
    ws_mocha.column_dimensions['B'].width = 50
    ws_mocha.column_dimensions['C'].width = 12
    ws_mocha.column_dimensions['D'].width = 14
    ws_mocha.column_dimensions['E'].width = 14
    ws_mocha.column_dimensions['F'].width = 10
    ws_mocha.column_dimensions['G'].width = 10
    
    # Tab 3: PL-Unit Tests (with suite and stats structure)
    ws_plunit = wb.create_sheet("PL-Unit Tests")
    for col, header in enumerate(headers, 1):
        cell = ws_plunit.cell(1, col, header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color='DDDDDD', end_color='DDDDDD', fill_type='solid')
    
    for row, test in enumerate(plunit['tests'], 2):
        full_title = test.get('fullTitle', '')
        # PL-Unit format is "module:test_name"
        if ':' in full_title:
            suite, test_name = full_title.split(':', 1)
        else:
            suite = ''
            test_name = full_title
        
        duration = test.get('duration', 0)
        
        ws_plunit.cell(row, 1, suite)
        ws_plunit.cell(row, 2, test_name)
        ws_plunit.cell(row, 3, duration)
        ws_plunit.cell(row, 4, duration)  # Average = single run for preview
        ws_plunit.cell(row, 5, 0)  # Std dev = 0 for single run
        ws_plunit.cell(row, 6, test.get('speed', 'n/a'))
        ws_plunit.cell(row, 7, test.get('state', 'unknown'))
    
    ws_plunit.column_dimensions['A'].width = 20
    ws_plunit.column_dimensions['B'].width = 50
    ws_plunit.column_dimensions['C'].width = 12
    ws_plunit.column_dimensions['D'].width = 14
    ws_plunit.column_dimensions['E'].width = 14
    ws_plunit.column_dimensions['F'].width = 10
    ws_plunit.column_dimensions['G'].width = 10
    
    # Save file
    timestamp = datetime.now().strftime('%Y-%m-%d')
    filename = f'benchmark_preview_{timestamp}.xlsx'
    filepath = results_dir / filename
    
    wb.save(filepath)
    
    print(f"✅ Excel report generated: {filepath}")
    print(f"\nSummary:")
    print(f"  Mocha:   {mocha['stats']['tests']} tests in {mocha['stats']['duration']}ms")
    print(f"  PL-Unit: {plunit['stats']['tests']} tests in {plunit['stats']['duration']}ms")

if __name__ == '__main__':
    main()
