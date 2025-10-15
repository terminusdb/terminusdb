#!/usr/bin/env python3
# Generate branch comparison Excel with statistics

import json
import sys
from datetime import datetime
from pathlib import Path
import statistics
import glob

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
except ImportError:
    print("❌ openpyxl not installed. Install with: pip3 install openpyxl")
    sys.exit(1)

def load_runs(runs_dir, test_type, suite_name):
    """Load all JSON files for a test type and suite"""
    runs = []
    pattern = f'{test_type}_{suite_name}_run_*.json'
    files = sorted(glob.glob(str(runs_dir / pattern)))
    
    for filepath in files:
        try:
            with open(filepath, 'r') as f:
                runs.append(json.load(f))
        except Exception as e:
            print(f"Warning: Error loading {filepath}: {e}")
    
    return runs

def calculate_test_stats(runs):
    """Calculate statistics for each test across runs"""
    test_data = {}
    
    for run in runs:
        for test in run.get('tests', []):
            full_title = test['fullTitle']
            if full_title not in test_data:
                test_data[full_title] = {
                    'durations': [],
                    'speed': test.get('speed', 'n/a'),
                    'state': test.get('state', 'unknown'),
                    'test': test
                }
            test_data[full_title]['durations'].append(test.get('duration', 0))
    
    results = []
    for full_title, data in sorted(test_data.items()):
        durations = data['durations']
        results.append({
            'fullTitle': full_title,
            'test': data['test'],
            'runs': durations,
            'average': statistics.mean(durations) if durations else 0,
            'stddev': statistics.stdev(durations) if len(durations) > 1 else 0,
            'min': min(durations) if durations else 0,
            'max': max(durations) if durations else 0,
            'speed': data['speed'],
            'state': data['state']
        })
    
    return results

def calculate_suite_stats(runs):
    """Calculate overall statistics for the test suite"""
    if not runs:
        return None
    
    durations = [run['stats']['duration'] for run in runs]
    return {
        'tests': runs[0]['stats']['tests'],
        'passes': runs[0]['stats']['passes'],
        'failures': runs[0]['stats']['failures'],
        'suites': runs[0]['stats']['suites'],
        'runs': durations,
        'average': statistics.mean(durations),
        'stddev': statistics.stdev(durations) if len(durations) > 1 else 0,
        'min': min(durations),
        'max': max(durations)
    }

def create_test_sheet(wb, sheet_name, test_stats, num_runs):
    """Create a test details sheet"""
    ws = wb.create_sheet(sheet_name)
    
    # Headers
    headers = ['Suite', 'Test Name']
    for i in range(1, num_runs + 1):
        headers.append(f'Run {i}')
    headers.extend(['Avg', 'StdDev', 'Min', 'Max', 'Speed', 'Status'])
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(1, col, header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color='DDDDDD', end_color='DDDDDD', fill_type='solid')
    
    for row, test_stat in enumerate(test_stats, 2):
        full_title = test_stat['fullTitle']
        
        # Parse suite and test name
        if ':' in full_title:
            suite, test_name = full_title.split(':', 1)
        else:
            parts = full_title.split(' ', 1)
            suite = parts[0] if parts else ''
            test_name = parts[1] if len(parts) > 1 else full_title
        
        ws.cell(row, 1, suite)
        ws.cell(row, 2, test_name)
        
        col = 3
        for duration in test_stat['runs']:
            ws.cell(row, col, round(duration))
            col += 1
        
        ws.cell(row, col, round(test_stat['average']))
        ws.cell(row, col + 1, round(test_stat['stddev'], 2))
        ws.cell(row, col + 2, round(test_stat['min']))
        ws.cell(row, col + 3, round(test_stat['max']))
        ws.cell(row, col + 4, test_stat['speed'])
        ws.cell(row, col + 5, test_stat['state'])
    
    # Column widths
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 50
    for i in range(3, 3 + num_runs + 6):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = 10

def create_comparison_overview(wb, branch1_name, branch2_name, branch1_data, branch2_data):
    """Create comparison overview sheet"""
    ws = wb.active
    ws.title = "Comparison"
    
    # Title
    ws.merge_cells('A1:K1')
    title_cell = ws['A1']
    title_cell.value = f"Branch Comparison: {branch1_name} vs {branch2_name}"
    title_cell.font = Font(bold=True, size=14)
    title_cell.alignment = Alignment(horizontal='center')
    
    # Headers
    row = 3
    headers = ['Test Suite', 'Tests', 'Branch', 'Avg (ms)', 'StdDev', 'Min', 'Max', 'Diff (ms)', 'Diff (%)', 'Faster']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row, col, header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color='DDDDDD', end_color='DDDDDD', fill_type='solid')
    
    row = 4
    
    # Group results by test suite
    all_suites = {}
    
    for suite_name, stats in branch1_data.items():
        if stats:
            all_suites[suite_name] = {'branch1': stats, 'branch2': None}
    
    for suite_name, stats in branch2_data.items():
        if stats:
            if suite_name in all_suites:
                all_suites[suite_name]['branch2'] = stats
            else:
                all_suites[suite_name] = {'branch1': None, 'branch2': stats}
    
    # Write comparison data
    for suite_name in sorted(all_suites.keys()):
        b1_stats = all_suites[suite_name]['branch1']
        b2_stats = all_suites[suite_name]['branch2']
        
        if b1_stats and b2_stats:
            # Both branches have this suite
            test_count = b1_stats['tests']
            
            # Branch 1 row
            ws.cell(row, 1, suite_name)
            ws.cell(row, 2, test_count)
            ws.cell(row, 3, branch1_name)
            ws.cell(row, 4, round(b1_stats['average']))
            ws.cell(row, 5, round(b1_stats['stddev'], 2))
            ws.cell(row, 6, round(b1_stats['min']))
            ws.cell(row, 7, round(b1_stats['max']))
            
            # Branch 2 row
            row += 1
            ws.cell(row, 1, '')
            ws.cell(row, 2, b2_stats['tests'])
            ws.cell(row, 3, branch2_name)
            ws.cell(row, 4, round(b2_stats['average']))
            ws.cell(row, 5, round(b2_stats['stddev'], 2))
            ws.cell(row, 6, round(b2_stats['min']))
            ws.cell(row, 7, round(b2_stats['max']))
            
            # Calculate difference
            diff_ms = b2_stats['average'] - b1_stats['average']
            diff_pct = (diff_ms / b1_stats['average'] * 100) if b1_stats['average'] > 0 else 0
            faster = branch1_name if diff_ms > 0 else branch2_name if diff_ms < 0 else 'Same'
            
            ws.cell(row, 8, round(diff_ms))
            ws.cell(row, 9, f"{diff_pct:.1f}%")
            ws.cell(row, 10, faster)
            
            # Color code
            if faster == branch1_name:
                ws.cell(row, 10).fill = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')
            elif faster == branch2_name:
                ws.cell(row, 10).fill = PatternFill(start_color='FFB6C1', end_color='FFB6C1', fill_type='solid')
            
            row += 1
        elif b1_stats:
            # Only in branch 1
            ws.cell(row, 1, suite_name)
            ws.cell(row, 2, b1_stats['tests'])
            ws.cell(row, 3, branch1_name)
            ws.cell(row, 4, round(b1_stats['average']))
            ws.cell(row, 10, 'Only in branch 1')
            row += 1
        elif b2_stats:
            # Only in branch 2
            ws.cell(row, 1, suite_name)
            ws.cell(row, 2, b2_stats['tests'])
            ws.cell(row, 3, branch2_name)
            ws.cell(row, 4, round(b2_stats['average']))
            ws.cell(row, 10, 'Only in branch 2')
            row += 1
    
    # Column widths
    ws.column_dimensions['A'].width = 25
    for col in ['B', 'C']:
        ws.column_dimensions[col].width = 12
    for col in ['D', 'E', 'F', 'G', 'H']:
        ws.column_dimensions[col].width = 10
    ws.column_dimensions['I'].width = 10
    ws.column_dimensions['J'].width = 15

def main():
    if len(sys.argv) < 4:
        print("Usage: benchmark_compare_excel.py <branch1> <branch2> <num_runs>")
        sys.exit(1)
    
    branch1_name = sys.argv[1]
    branch2_name = sys.argv[2]
    num_runs = int(sys.argv[3])
    
    base_dir = Path(__file__).parent
    branch1_dir = base_dir / 'benchmark_results' / 'branch1'
    branch2_dir = base_dir / 'benchmark_results' / 'branch2'
    
    print("Loading test results...")
    
    # Discover all test suites from files
    branch1_files = list(branch1_dir.glob('*_run_1.json'))
    branch2_files = list(branch2_dir.glob('*_run_1.json'))
    
    all_suites = set()
    for f in branch1_files:
        # Extract suite name from filename (e.g., mocha_capabilities_run_1.json -> capabilities)
        parts = f.stem.split('_')
        if len(parts) >= 3:
            test_type = parts[0]
            suite_name = '_'.join(parts[1:-2])  # Everything between type and "run_N"
            all_suites.add((test_type, suite_name))
    
    for f in branch2_files:
        parts = f.stem.split('_')
        if len(parts) >= 3:
            test_type = parts[0]
            suite_name = '_'.join(parts[1:-2])
            all_suites.add((test_type, suite_name))
    
    print(f"Found {len(all_suites)} test suites to compare")
    
    # Load and calculate stats for all suites
    branch1_data = {}
    branch2_data = {}
    
    wb = openpyxl.Workbook()
    
    for test_type, suite_name in sorted(all_suites):
        print(f"Processing {test_type}/{suite_name}...")
        
        # Branch 1
        b1_runs = load_runs(branch1_dir, test_type, suite_name)
        if b1_runs:
            b1_stats = calculate_suite_stats(b1_runs)
            b1_tests = calculate_test_stats(b1_runs)
            branch1_data[f'{test_type}_{suite_name}'] = b1_stats
            
            # Create sheet for branch 1
            sheet_name = f'{test_type}_{suite_name}_b1'[:31]  # Excel limit
            create_test_sheet(wb, sheet_name, b1_tests, num_runs)
        
        # Branch 2
        b2_runs = load_runs(branch2_dir, test_type, suite_name)
        if b2_runs:
            b2_stats = calculate_suite_stats(b2_runs)
            b2_tests = calculate_test_stats(b2_runs)
            branch2_data[f'{test_type}_{suite_name}'] = b2_stats
            
            # Create sheet for branch 2
            sheet_name = f'{test_type}_{suite_name}_b2'[:31]  # Excel limit
            create_test_sheet(wb, sheet_name, b2_tests, num_runs)
    
    # Create comparison overview
    create_comparison_overview(wb, branch1_name, branch2_name, branch1_data, branch2_data)
    
    # Save file
    timestamp = datetime.now().strftime('%Y-%m-%d_%H%M%S')
    filename = f'benchmark_comparison_{branch1_name}_vs_{branch2_name}_{timestamp}.xlsx'
    # Sanitize filename
    filename = filename.replace('/', '_').replace(' ', '_')
    filepath = base_dir / 'benchmark_results' / filename
    
    wb.save(filepath)
    
    print(f"\n✅ Comparison report generated: {filepath}")
    print(f"\nSheets created:")
    print(f"  - Comparison (overview)")
    for sheet in wb.sheetnames[1:]:
        print(f"  - {sheet}")

if __name__ == '__main__':
    main()
